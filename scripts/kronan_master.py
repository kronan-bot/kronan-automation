"""
KrÃ³nan Master Sales Tracker
Run: python3 kronan_master.py <new_report.xlsx>
- Appends new daily data to master file (no duplicates)
- Regenerates monthly summaries automatically
"""

import openpyxl, datetime, re
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from collections import defaultdict
import sys, os
try:
    from openpyxl.cell.cell import MergedCell
except ImportError:
    from openpyxl.cell import MergedCell

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_BASE = os.environ.get('KRONAN_BASE')
MASTER = os.path.join(_BASE, 'KrÃ³nan_Master_SkrÃ¡.xlsx') if _BASE else os.path.join(os.path.expanduser("~"), "Documents", "KrÃ³nan", "KrÃ³nan_Master_SkrÃ¡.xlsx")

# ââ Styles ââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
BLUE = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
DARK = PatternFill("solid", start_color="1A5276", end_color="1A5276")
GREEN = PatternFill("solid", start_color="1E8449", end_color="1E8449")
MHDR = PatternFill("solid", start_color="117A65", end_color="117A65")
ALT = PatternFill("solid", start_color="EBF5FB", end_color="EBF5FB")
ALT2 = PatternFill("solid", start_color="E9F7EF", end_color="E9F7EF")
WHITE = PatternFill("solid", start_color="FFFFFF", end_color="FFFFFF")
TITL = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
thin = Side(style="thin", color="BDC3C7")
brd = Border(left=thin, right=thin, top=thin, bottom=thin)
hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
df = Font(name="Arial", size=10)
tf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
titf = Font(name="Arial", bold=True, color="1F4E79", size=13)
secf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
ctr = Alignment(horizontal="center", vertical="center")
lft = Alignment(horizontal="left", vertical="center")
rgt = Alignment(horizontal="right", vertical="center")

def style_header(cell, text, fill=BLUE):
    cell.value = cell.value if text is None else text
    cell.fill = fill; cell.font = hf; cell.alignment = ctr; cell.border = brd

def style_data(cell, val, fmt=None, aln=None, fill=WHITE):
    cell.value = val
    cell.font = df; cell.fill = fill; cell.border = brd
    cell.alignment = aln or lft
    if fmt: cell.number_format = fmt

def style_total(cell, val, fmt=None, aln=None):
    cell.value = val; cell.fill = DARK; cell.font = tf; cell.border = brd
    cell.alignment = aln or lft
    if fmt: cell.number_format = fmt

def style_month_hdr(cell, val):
    cell.value = val; cell.fill = GREEN; cell.font = secf
    cell.alignment = lft; cell.border = brd

# ââ Date parsing helper ââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def _try_parse_date(val):
    """Return a datetime object if val is or looks like a date; else None."""
    if val is None:
        return None
    if hasattr(val, 'strftime'):  # datetime.datetime or datetime.date from openpyxl
        return val
    if isinstance(val, str):
        vs = val.strip()
        # SQL Server format: "Jun 14 2026 15:53:20:773"
        m = re.match(r'([A-Za-z]{3})\s+(\d{1,2})\s+(\d{4})', vs)
        if m:
            try:
                return datetime.datetime.strptime(
                    '{} {:02d} {}'.format(m.group(1), int(m.group(2)), m.group(3)),
                    "%b %d %Y")
            except Exception:
                pass
        # ISO date: "2026-06-14"
        m2 = re.match(r'(\d{4}-\d{2}-\d{2})', vs)
        if m2:
            try:
                return datetime.datetime.strptime(m2.group(1), "%Y-%m-%d")
            except Exception:
                pass
    return None

# ââ Clean corrupted master rows ââââââââââââââââââââââââââââââââââââââââââââââ
def clean_master_dates(wb):
    """Remove rows from daily sheets where date column is not a valid date."""
    total = 0
    for sname in ["Dagleg - Verslanir", "Dagleg - VaraÃVerslun", "Dagleg - VÃ¶rur"]:
        if sname not in wb.sheetnames:
            continue
        ws = wb[sname]
        bad = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            if _try_parse_date(row[0]) is None:
                bad.append((i, str(row[0])[:60]))
        if bad:
            print(f"   Cleaning {len(bad)} bad row(s) from {sname}: {[v for _,v in bad]}")
            for row_idx, _ in reversed(bad):
                ws.delete_rows(row_idx)
            total += len(bad)
    if total == 0:
        print("   No corrupted date rows found in master")
    return total

# ââ Read incoming report âââââââââââââââââââââââââââââââââââââââââââââââââââââ
def read_report(path):
    wb = openpyxl.load_workbook(path)
    # Stores – try canonical name, fall back to first sheet
    try:
        ws_s = wb['Verslanir']
    except KeyError:
        ws_s = wb.worksheets[0]
        print(f" ⚠ Sheet 'Verslanir' not found, using first sheet: {ws_s.title}")

    # Debug: print column structure to diagnose format changes
    max_col = min(ws_s.max_column, 14)
    hdr = [ws_s.cell(1, c).value for c in range(1, max_col + 1)]
    print(f" Sheet headers (row 1): {hdr}")
    for r in ws_s.iter_rows(min_row=2, max_row=2, values_only=True):
        print(f" First data row  (row 2): {list(r)[:max_col]}")

    store_rows         = defaultdict(lambda: [0.0, 0])
    item_rows          = defaultdict(lambda: [0.0, 0])
    store_product_rows = defaultdict(lambda: defaultdict(lambda: [0.0, 0, '']))
    date = None

    for row in ws_s.iter_rows(min_row=2, values_only=True):
        if len(row) < 9: continue
        # Actual column mapping (0-indexed) - 9-column Kronan format:
        # 0=Dags (Date), 1=Kedja (Chain), 2=EAN, 3=Heiti verslunar (Store),
        # 4=Voru Nr (Product No), 5=Vara (Product Name), 6=Vendor No,
        # 7=Sala (Sales ISK), 8=Magn (Quantity)
        d_raw = row[0]   # Date
        store = row[3]   # Store Name
        pnr   = row[4]   # Product No (internal)
        prod  = row[5]   # Product Name
        sale  = row[7]   # Sales (ISK)
        qty   = row[8]   # Quantity

        if not store or not prod: continue
        if sale is None: sale = 0.0

        # Extract date from first valid row
        if date is None:
            d_parsed = _try_parse_date(d_raw)
            if d_parsed is None:
                # Fallback: scan all columns for a date-like value
                for ci, cv in enumerate(row):
                    d_parsed = _try_parse_date(cv)
                    if d_parsed:
                        print(f" Date found via row-scan (col {ci}): {d_parsed}")
                        break
            if d_parsed:
                date = d_parsed

        store_rows[store][0]               += float(sale)
        store_rows[store][1]               += int(qty or 0)
        item_rows[prod][0]                 += float(sale)
        item_rows[prod][1]                 += int(qty or 0)
        store_product_rows[store][prod][0] += float(sale)
        store_product_rows[store][prod][1] += int(qty or 0)
        store_product_rows[store][prod][2]  = str(pnr or '')

    print(f" Date: {date}, Stores: {len(store_rows)}, Products: {len(item_rows)}")
    return date, store_rows, item_rows, store_product_rows

def load_or_create():
    if os.path.exists(MASTER):
        return load_workbook(MASTER)
    wb = Workbook()
    # Sheet 1: Daily stores
    ws = wb.active; ws.title = "Dagleg - Verslanir"
    for col, (h, w) in enumerate(zip(
        ["Dags","MÃ¡nuÃ°ur","Verslun","Sala (kr)","Magn","% Dagsins"],
        [14, 12, 32, 18, 10, 12]), 1):
        c = ws.cell(1, col); style_header(c, h)
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "A2"
    # Sheet 2: Monthly stores
    ws2 = wb.create_sheet("MÃ¡naÃ°arleg - Verslanir")
    for col, (h, w) in enumerate(zip(
        ["MÃ¡nuÃ°ur","Verslun","Sala (kr)","Magn","% MÃ¡naÃ°arins"],
        [14, 32, 18, 10, 16]), 1):
        c = ws2.cell(1, col); style_header(c, h, MHDR)
        ws2.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws2.row_dimensions[1].height = 20
    ws2.freeze_panes = "A2"
    # Sheet 3: Daily products
    ws3 = wb.create_sheet("Dagleg - VÃ¶rur")
    for col, (h, w) in enumerate(zip(
        ["Dags","MÃ¡nuÃ°ur","Vara","Sala (kr)","Magn","% Dagsins"],
        [14, 12, 44, 18, 10, 12]), 1):
        c = ws3.cell(1, col); style_header(c, h)
        ws3.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws3.row_dimensions[1].height = 20
    ws3.freeze_panes = "A2"
    # Sheet 4: Monthly products
    ws4 = wb.create_sheet("MÃ¡naÃ°arleg - VÃ¶rur")
    for col, (h, w) in enumerate(zip(
        ["MÃ¡nuÃ°ur","Vara","Sala (kr)","Magn","% MÃ¡naÃ°arins"],
        [14, 44, 18, 10, 16]), 1):
        c = ws4.cell(1, col); style_header(c, h, MHDR)
        ws4.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws4.row_dimensions[1].height = 20
    ws4.freeze_panes = "A2"
    # Sheet 5: Daily storeÃproduct breakdown
    ws5 = wb.create_sheet("Dagleg - VaraÃVerslun")
    for col, (h, w) in enumerate(zip(
        ["Dags","MÃ¡nuÃ°ur","Verslun","Vara","Sala (kr)","Magn"],
        [14, 12, 32, 44, 18, 10]), 1):
        c = ws5.cell(1, col); style_header(c, h)
        ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
    ws5.row_dimensions[1].height = 20
    ws5.freeze_panes = "A2"
    return wb

# ââ Append daily data ââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def append_daily(wb, date, store_rows, item_rows, store_product_rows=None):
    date_val = date.date() if hasattr(date, 'date') else date
    month_str = date.strftime("%Y-%m") if hasattr(date, 'strftime') else str(date)[:7]

    # --- Stores sheet ---
    ws = wb["Dagleg - Verslanir"]
    # Check for duplicates
    existing_dates = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            d = row[0].date() if hasattr(row[0], 'date') else row[0]
            existing_dates.add(d)
    if date_val in existing_dates:
        print(f" â  {date_val} already in Dagleg - Verslanir, skipping.")
    else:
        store_total = sum(v[0] for v in store_rows.values())
        stores_sorted = sorted(store_rows.items(), key=lambda x: -x[1][0])
        r = ws.max_row + 1
        for store, (sale, qty) in stores_sorted:
            fill = ALT if (r % 2 == 0) else WHITE
            pct = sale / store_total if store_total else 0
            for col, (val, fmt, aln) in enumerate([
                (date, "DD.MM.YYYY", ctr),
                (month_str, None, ctr),
                (store, None, lft),
                (sale, '#,##0.00', rgt),
                (qty, '#,##0', ctr),
                (pct, '0.0%', ctr),
            ], 1):
                style_data(ws.cell(r, col), val, fmt, aln, fill)
            ws.row_dimensions[r].height = 17
            r += 1
        print(f" â Appended {len(stores_sorted)} store rows for {date_val}")

    # --- Products sheet ---
    ws3 = wb["Dagleg - VÃ¶rur"]
    existing_dates3 = set()
    for row in ws3.iter_rows(min_row=2, values_only=True):
        if row[0]:
            d = row[0].date() if hasattr(row[0], 'date') else row[0]
            existing_dates3.add(d)
    if date_val in existing_dates3:
        print(f" â  {date_val} already in Dagleg - VÃ¶rur, skipping.")
    else:
        item_total = sum(v[0] for v in item_rows.values())
        items_sorted = sorted(item_rows.items(), key=lambda x: -x[1][0])
        r = ws3.max_row + 1
        for prod, (sale, qty) in items_sorted:
            fill = ALT if (r % 2 == 0) else WHITE
            pct = sale / item_total if item_total else 0
            for col, (val, fmt, aln) in enumerate([
                (date, "DD.MM.YYYY", ctr),
                (month_str, None, ctr),
                (prod, None, lft),
                (sale, '#,##0.00', rgt),
                (qty, '#,##0', ctr),
                (pct, '0.0%', ctr),
            ], 1):
                style_data(ws3.cell(r, col), val, fmt, aln, fill)
            ws3.row_dimensions[r].height = 17
            r += 1
        print(f" â Appended {len(items_sorted)} product rows for {date_val}")

    # --- StoreÃProduct sheet ---
    if "Dagleg - VaraÃVerslun" not in wb.sheetnames:
        ws5 = wb.create_sheet("Dagleg - VaraÃVerslun")
        for col, (h, w) in enumerate(zip(
            ["Dags","MÃ¡nuÃ°ur","Verslun","Vara","Sala (kr)","Magn","VÃ¶runÃºmer"],
            [14, 12, 32, 44, 18, 10, 14]), 1):
            c = ws5.cell(1, col); style_header(c, h)
            ws5.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w
        ws5.row_dimensions[1].height = 20
        ws5.freeze_panes = "A2"
    ws5 = wb["Dagleg - VaraÃVerslun"]
    existing_dates5 = set()
    for row in ws5.iter_rows(min_row=2, values_only=True):
        if row[0]:
            d = row[0].date() if hasattr(row[0], 'date') else row[0]
            existing_dates5.add(d)
    if date_val in existing_dates5:
        print(f" â  {date_val} already in Dagleg - VaraÃVerslun, skipping.")
    else:
        r = ws5.max_row + 1
        count = 0
        for store in sorted(store_product_rows.keys()):
            prods = sorted(store_product_rows[store].items(), key=lambda x: -x[1][0])
            for prod, (sale, qty, pnr) in prods:
                fill = ALT if (r % 2 == 0) else WHITE
                for col, (val, fmt, aln) in enumerate([
                    (date, "DD.MM.YYYY", ctr),
                    (month_str, None, ctr),
                    (store, None, lft),
                    (prod, None, lft),
                    (sale, '#,##0.00', rgt),
                    (qty, '#,##0', ctr),
                    (pnr, None, ctr),
                ], 1):
                    style_data(ws5.cell(r, col), val, fmt, aln, fill)
                ws5.row_dimensions[r].height = 17
                r += 1; count += 1
        print(f" â Appended {count} storeÃproduct rows for {date_val}")

# ââ Rebuild monthly summaries ââââââââââââââââââââââââââââââââââââââââââââââââ
def rebuild_monthly(wb):
    # --- Monthly Stores ---
    ws_daily = wb["Dagleg - Verslanir"]
    monthly_stores = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
    for row in ws_daily.iter_rows(min_row=2, values_only=True):
        d, month, store, sale, qty, pct = row
        if month and store and sale:
            monthly_stores[month][store][0] += sale
            monthly_stores[month][store][1] += int(qty or 0)

    ws_m = wb["MÃ¡naÃ°arleg - Verslanir"]
    # Unmerge all, then clear
    for merge in list(ws_m.merged_cells.ranges):
        ws_m.unmerge_cells(str(merge))
    for row in ws_m.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell, MergedCell): continue
            cell.value = None; cell.fill = WHITE; cell.border = Border()

    r = 2
    for month in sorted(monthly_stores.keys()):
        stores = sorted(monthly_stores[month].items(), key=lambda x: -x[1][0])
        month_total = sum(v[0] for _, v in stores)
        # Month header
        ws_m.merge_cells(f"A{r}:E{r}")
        c = ws_m.cell(r, 1)
        style_month_hdr(c, f" {month}")
        ws_m.row_dimensions[r].height = 18
        r += 1
        for store, (sale, qty) in stores:
            fill = ALT if (r % 2 == 0) else WHITE
            pct = sale / month_total if month_total else 0
            for col, (val, fmt, aln) in enumerate([
                (month, None, ctr), (store, None, lft),
                (sale, '#,##0.00', rgt), (qty, '#,##0', ctr), (pct, '0.0%', ctr)
            ], 1):
                style_data(ws_m.cell(r, col), val, fmt, aln, fill)
            ws_m.row_dimensions[r].height = 17
            r += 1
        # Month total row
        for col, (val, fmt, aln) in enumerate([
            (month, None, ctr), (f"HEILD {month}", None, lft),
            (month_total, '#,##0.00', rgt), (sum(v[1] for _,v in stores), '#,##0', ctr),
            ("100.0%", '0.0%', ctr)
        ], 1):
            style_total(ws_m.cell(r, col), val, fmt, aln)
        ws_m.row_dimensions[r].height = 18
        r += 1
    print(f" â Monthly stores summary rebuilt ({len(monthly_stores)} months)")

    # --- Monthly Products ---
    ws_daily3 = wb["Dagleg - VÃ¶rur"]
    monthly_items = defaultdict(lambda: defaultdict(lambda: [0.0, 0]))
    for row in ws_daily3.iter_rows(min_row=2, values_only=True):
        d, month, prod, sale, qty, pct = row
        if month and prod and sale:
            monthly_items[month][prod][0] += sale
            monthly_items[month][prod][1] += int(qty or 0)

    ws_m4 = wb["MÃ¡naÃ°arleg - VÃ¶rur"]
    # Unmerge all, then clear
    for merge in list(ws_m4.merged_cells.ranges):
        ws_m4.unmerge_cells(str(merge))
    for row in ws_m4.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell, MergedCell): continue
            cell.value = None; cell.fill = WHITE; cell.border = Border()

    r = 2
    for month in sorted(monthly_items.keys()):
        items = sorted(monthly_items[month].items(), key=lambda x: -x[1][0])
        month_total = sum(v[0] for _, v in items)
        ws_m4.merge_cells(f"A{r}:E{r}")
        c = ws_m4.cell(r, 1)
        style_month_hdr(c, f" {month}")
        ws_m4.row_dimensions[r].height = 18
        r += 1
        for prod, (sale, qty) in items:
            fill = ALT2 if (r % 2 == 0) else WHITE
            pct = sale / month_total if month_total else 0
            for col, (val, fmt, aln) in enumerate([
                (month, None, ctr), (prod, None, lft),
                (sale, '#,##0.00', rgt), (qty, '#,##0', ctr), (pct, '0.0%', ctr)
            ], 1):
                style_data(ws_m4.cell(r, col), val, fmt, aln, fill)
            ws_m4.row_dimensions[r].height = 17
            r += 1
        for col, (val, fmt, aln) in enumerate([
            (month, None, ctr), (f"HEILD {month}", None, lft),
            (month_total, '#,##0.00', rgt), (sum(v[1] for _,v in items), '#,##0', ctr),
            ("100.0%", '0.0%', ctr)
        ], 1):
            style_total(ws_m4.cell(r, col), val, fmt, aln)
        ws_m4.row_dimensions[r].height = 18
        r += 1
    print(f" â Monthly products summary rebuilt ({len(monthly_items)} months)")

# ââ Main âââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââââ
def run(report_path):
    print(f"\n\U0001f4c2 Reading report: {report_path}")
    date, store_rows, item_rows, store_product_rows = read_report(report_path)
    print(f" Date: {date}, Stores: {len(store_rows)}, Products: {len(item_rows)}")
    print(f"\n\U0001f4d2 Loading master file...")
    wb = load_or_create()
    print(f"\n\U0001f9f9 Checking master for corrupted rows...")
    clean_master_dates(wb)
    print(f"\nâ Appending daily data...")
    append_daily(wb, date, store_rows, item_rows, store_product_rows)
    print(f"\n\U0001f4c5 Rebuilding monthly summaries...")
    rebuild_monthly(wb)
    wb.save(MASTER)
    print(f"\nâ Master file saved: {MASTER}")

if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else os.path.join(SCRIPT_DIR, 'KrÃ³nan sÃ¶luskÃ½rsla.xlsx')
    run(path)
