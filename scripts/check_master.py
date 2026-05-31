import openpyxl, os, sys
m = os.environ.get('KRONAN_BASE','data') + '/Kronan_Master_Skra.xlsx'
# Try both filename variants
for name in ['Krónan_Master_Skrá.xlsx', 'Kronan_Master_Skra.xlsx']:
    p = os.path.join(os.environ.get('KRONAN_BASE','data'), name)
    if os.path.exists(p):
        m = p
        break

print(f'Master path: {m}')
print(f'Master size: {os.path.getsize(m):,} bytes')
wb = openpyxl.load_workbook(m, data_only=True)
print(f'Sheets: {wb.sheetnames}')
ws = wb['Dagleg - Verslanir']
dates = sorted(set(
    r[0].strftime('%Y-%m-%d') for r in ws.iter_rows(min_row=2, values_only=True)
    if r[0] and hasattr(r[0], 'strftime')
))
print(f'Dates: {len(dates)}, first={dates[0] if dates else None}, last={dates[-1] if dates else None}')
# Write to file for git commit
out = os.path.join(os.environ.get('KRONAN_BASE','data'), 'master_check.txt')
with open(out, 'w') as f:
    f.write(f'size={os.path.getsize(m)}, dates={len(dates)}, first={dates[0] if dates else None}, last={dates[-1] if dates else None}')
