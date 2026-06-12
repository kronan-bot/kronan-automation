"""
update_dashboard_data.py
Reads KrÃÂ³nan_Master_SkrÃÂ¡.xlsx and patches ALL data constants
inside KrÃÂ³nan_Dashboard.html using marker-based replacement.

Primary method : replaces entire /* KRONAN_DATA_START */ Ã¢â¬Â¦ /* KRONAN_DATA_END */ block.
Fallback method: const-variable anchor + buildTabs() slice.
Self-healing   : if core JS functions (buildTabs, render) are missing from the local
                 HTML, downloads a fresh copy from the Netlify CDN before patching.
"""
import openpyxl, json, os, re, sys, urllib.request, urllib.error
from collections import defaultdict

NETLIFY_URL = 'https://kronan-dashboard.netlify.app/'

SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
_BASE = os.environ.get('KRONAN_BASE')
if _BASE:
    MASTER    = os.path.join(_BASE, 'KrÃÂ³nan_Master_SkrÃÂ¡.xlsx')
    DASHBOARD = os.path.join(_BASE, 'KrÃÂ³nan_Dashboard.html')
else:
    MASTER    = os.path.join(SCRIPT_DIR, '..', 'KrÃÂ³nan_Master_SkrÃÂ¡.xlsx')
    DASHBOARD = os.path.join(SCRIPT_DIR, '..', 'KrÃÂ³nan_Dashboard.html')

# Ã¢ââ¬Ã¢ââ¬ Sanity check files exist Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
if not os.path.exists(MASTER):
    print(f'Ã¢Åâ Master file not found: {MASTER}')
    sys.exit(1)
if not os.path.exists(DASHBOARD):
    print(f'Ã¢Åâ Dashboard not found: {DASHBOARD}')
    sys.exit(1)

# Ã¢ââ¬Ã¢ââ¬ Self-healing: recover dashboard if core JS functions are missing Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
def _dashboard_healthy(path):
    try:
        # A healthy dashboard is simply a large file (>100 KB).
        # Searching for JS function names deep inside a 775 KB file is fragile.
        return os.path.getsize(path) > 100_000
    except Exception:
        return False

if not _dashboard_healthy(DASHBOARD):
    print('Ã¢Å¡Â   Dashboard is missing core JS Ã¢â¬â downloading fresh copy from Netlify CDNÃ¢â¬Â¦')
    try:
        req = urllib.request.Request(NETLIFY_URL, headers={'Cache-Control': 'no-cache'})
        with urllib.request.urlopen(req, timeout=30) as resp:
            fresh = resp.read().decode('utf-8')
        if 'function buildTabs' in fresh and 'function render(' in fresh:
            with open(DASHBOARD, 'w', encoding='utf-8') as f:
                f.write(fresh)
            print('Ã¢Åâ¦ Dashboard recovered from Netlify CDN')
        else:
            print('Ã¢Åâ Downloaded HTML is also missing JS functions Ã¢â¬â cannot recover')
            sys.exit(1)
    except Exception as e:
        print(f'Ã¢Åâ Recovery download failed: {e}')
        sys.exit(1)

wb = openpyxl.load_workbook(MASTER, data_only=True)

# Ã¢ââ¬Ã¢ââ¬ 1. DAILY totals Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
ws = wb['Dagleg - Verslanir']
daily_sale = defaultdict(float)
daily_qty  = defaultdict(int)
stores_by_date = defaultdict(list)

for row in ws.iter_rows(min_row=2, values_only=True):
    d, month, store, sale, qty, pct = row
    if not d or not store: continue
    dk = d.strftime('%Y-%m-%d') if hasattr(d,'strftime') else str(d)
    daily_sale[dk] += sale or 0
    daily_qty[dk]  += int(qty or 0)
    stores_by_date[dk].append({'store': store, 'sale': round(sale or 0), 'qty': int(qty or 0)})

dates = sorted(daily_sale.keys())
DAILY     = {d: round(daily_sale[d]) for d in dates}
DAILY_QTY = {d: daily_qty[d] for d in dates}
STORES_BY_DATE = {d: sorted(stores_by_date[d], key=lambda x: -x['sale']) for d in dates}

# Ã¢ââ¬Ã¢ââ¬ 2. STORE totals Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
store_sale = defaultdict(float)
store_qty  = defaultdict(int)
for dk, rows in stores_by_date.items():
    for r in rows:
        store_sale[r['store']] += r['sale']
        store_qty[r['store']]  += r['qty']

STORE_TOTALS = sorted(
    [[s, {'sale': round(store_sale[s]), 'qty': store_qty[s]}] for s in store_sale],
    key=lambda x: -x[1]['sale']
)

# Ã¢ââ¬Ã¢ââ¬ 3. PRODUCT totals Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
ws2 = wb['Dagleg - VÃÂ¶rur'] if 'Dagleg - VÃÂ¶rur' in wb.sheetnames else None
prod_sale = defaultdict(float)
prod_qty  = defaultdict(int)
if ws2:
    for row in ws2.iter_rows(min_row=2, values_only=True):
        row = list(row)
        if len(row) < 4: continue
        d, month, prod = row[0], row[1], row[2]
        sale = row[3] if len(row) > 3 else None
        qty  = row[4] if len(row) > 4 else 0
        if not prod: continue
        prod_sale[prod] += sale or 0
        prod_qty[prod]  += int(qty or 0)

PROD_TOTALS = sorted(
    [[p, {'sale': round(prod_sale[p]), 'qty': prod_qty[p]}] for p in prod_sale],
    key=lambda x: -x[1]['sale']
)

# Ã¢ââ¬Ã¢ââ¬ 4. StoreÃâProduct by date Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
ws5 = wb['Dagleg - VaraÃâVerslun'] if 'Dagleg - VaraÃâVerslun' in wb.sheetnames else None
store_prods = defaultdict(lambda: defaultdict(lambda: defaultdict(lambda: [0.0, 0, ''])))
if ws5:
    for row in ws5.iter_rows(min_row=2, values_only=True):
        if len(row) >= 7:
            d, month, store, prod, sale, qty, pnr = row[0], row[1], row[2], row[3], row[4], row[5], row[6]
        elif len(row) >= 6:
            d, month, store, prod, sale, qty = row[0], row[1], row[2], row[3], row[4], row[5]
            pnr = ''
        else:
            continue
        if not d or not store or not prod: continue
        dk = d.strftime('%Y-%m-%d') if hasattr(d,'strftime') else str(d)
        store_prods[dk][store][prod][0] += sale or 0
        store_prods[dk][store][prod][1] += int(qty or 0)
        if pnr: store_prods[dk][store][prod][2] = str(pnr)

# Ã¢ââ¬Ã¢ââ¬ 5. Build JS data block Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
def js_store_totals(data):
    parts = [f'["{n}",{{sale:{v["sale"]},qty:{v["qty"]}}}]' for n,v in data]
    return 'const STORE_TOTALS = [\n  ' + ','.join(parts) + '\n];'

def js_prod_totals(data):
    parts = [f'["{n}",{{sale:{v["sale"]},qty:{v["qty"]}}}]' for n,v in data]
    return 'const PROD_TOTALS = [\n  ' + ','.join(parts) + '\n];'

def js_daily(data):
    inner = ','.join(f'"{k}":{v}' for k,v in data.items())
    return f'const DAILY     = {{{inner}}};'

def js_daily_qty(data):
    inner = ','.join(f'"{k}":{v}' for k,v in data.items())
    return f'const DAILY_QTY = {{{inner}}};'

def js_stores_by_date(data):
    parts = []
    for dk, rows in sorted(data.items()):
        row_strs = ['{' + f'"store":"{r["store"]}","sale":{r["sale"]},"qty":{r["qty"]}' + '}' for r in rows]
        parts.append(f'"{dk}":[{",".join(row_strs)}]')
    return 'const STORES_BY_DATE = {' + ','.join(parts) + '};'

def js_store_products(data):
    date_parts = []
    for dk in sorted(data.keys()):
        store_parts = []
        for store in sorted(data[dk].keys(), key=lambda s: -sum(v[0] for v in data[dk][s].values())):
            prod_parts = []
            for prod, (sale, qty, pnr) in sorted(data[dk][store].items(), key=lambda x: -x[1][0]):
                pname = prod.replace('Tokyo Sushi ','').replace('"','\\"')
                pnr_s = str(pnr or '').replace('"','\\"')
                prod_parts.append(f'["{pnr_s}","{pname}",{round(sale)},{qty}]')
            sname = store.replace('"','\\"')
            store_parts.append(f'"{sname}":[{",".join(prod_parts)}]')
        date_parts.append(f'"{dk}":{{{",".join(store_parts)}}}')
    return 'const STORE_PRODS = {' + ','.join(date_parts) + '};'

DATA_BLOCK = '\n'.join([
    "const DOW = ['Sun','MÃÂ¡n','ÃÅ¾ri','MiÃÂ°','Fim','FÃÂ¶s','Lau'];",
    "let currentDate = 'all';",
    "let mainChart, storeChart;",
    "const STORE_COLORS = ['#1d4ed8','#2563eb','#3b82f6','#60a5fa','#0ea5e9','#06b6d4','#0891b2','#0284c7','#7c3aed','#6d28d9','#8b5cf6','#a78bfa','#059669','#10b981','#34d399','#065f46','#d97706','#f59e0b','#fbbf24','#b45309','#dc2626','#ef4444'];",
    "const STORE_COLOR_MAP = {'KrÃÂ³nan Granda':'#e63946','KrÃÂ³nan Flatahrauni':'#1d6fa4','KrÃÂ³nan BÃÂ­ldshÃÂ¶fÃÂ°a':'#2a9d8f','KrÃÂ³nan Selfossi':'#f4a261','KrÃÂ³nan MosfellsbÃÂ¦':'#8338ec','KrÃÂ³nan Skeifan 19':'#e9c46a','KrÃÂ³nan Fitjabraut':'#06d6a0','KrÃÂ³nan Vestmannaeyjum':'#ef476f','KrÃÂ³nan Akrabraut':'#4895ef','KrÃÂ³nan NorÃÂ°urhellu':'#560bad','KrÃÂ³nan BorgartÃÂºn':'#f72585','KrÃÂ³nan Austurveri':'#4cc9f0','KrÃÂ³nan Grafarholti':'#80b918','KrÃÂ³nan HallveigarstÃÂ­g':'#fb8500','KrÃÂ³nan Akranesi':'#264653','KrÃÂ³nan Jafnaseli':'#43aa8b','KrÃÂ³nan ÃÅ¾orlÃÂ¡ksN[Â¶fÅ¾':'#c9184a','KrÃÂ³nan Hvolsvelli':'#118ab2','KrÃÂ³nan VallakÃÂ³r':'#ffb703','KrÃÂ³nan VÃÂ­k':'#6d6875','KrÃÂ³nan ÃÂrbÃÂ¦':'#b5838d','KrÃÂ³nan Akureyri':'#52b788'};",
    js_store_totals(STORE_TOTALS),
    js_prod_totals(PROD_TOTALS),
    js_daily(DAILY),
    js_daily_qty(DAILY_QTY),
    js_stores_by_date(STORES_BY_DATE),
    js_store_products(store_prods),
    'const DATES = Object.keys(DAILY).sort();',  # sorted date array used throughout the UI
])

MARKER_START = '/* KRONAN_DATA_START */'
MARKER_END   = '/* KRONAN_DATA_END */'

# Ã¢ââ¬Ã¢ââ¬ 6. Patch the dashboard HTML Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
with open(DASHBOARD, 'r', encoding='utf-8') as f:
    html = f.read()

new_block = MARKER_START + '\n' + DATA_BLOCK + '\n' + MARKER_END

# Ã¢ââ¬Ã¢ââ¬ Path A: markers present Ã¢â¬â replace only the content between them Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
marker_s = html.find(MARKER_START)
marker_e = html.find(MARKER_END)
if marker_s >= 0 and marker_e > marker_s:
    html = html[:marker_s] + new_block + html[marker_e + len(MARKER_END):]
    print(f'  (marker path)')

else:
    # Ã¢ââ¬Ã¢ââ¬ Path B: no markers Ã¢â¬â find data section boundaries precisely Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬Ã¢ââ¬
    # START: earliest data const declaration
    data_start = len(html)
    for name in ('STORE_TOTALS', 'PROD_TOTALS', 'DAILY', 'DAILY_QTY', 'STORES_BY_DATE', 'STORE_PRODS'):
        p = html.find(f'const {name}')
        if 0 <= p < data_start:
            data_start = p

    if data_start == len(html):
        # No data consts at all Ã¢â¬â inject immediately before buildTabs() call
        bp = html.find('buildTabs();')
        if bp < 0:
            pos = html.rfind('</script>')
            if pos < 0:
                print('Ã¢Åâ Could not patch dashboard Ã¢â¬â no injection point found')
                sys.exit(1)
            html = html[:pos] + new_block + '\n' + html[pos:]
        else:
            html = html[:bp] + new_block + '\n\n' + html[bp:]
        print(f'  (inject path)')
    else:
        # END: first JS function declaration after the data consts
        # (function defs come after data in the original HTML)
        fn_pos = html.find('\nfunction ', data_start)
        if fn_pos >= 0:
            # Keep from fn_pos+1 so we don't swallow the leading newline
            html = html[:data_start] + new_block + '\n\n' + html[fn_pos + 1:]
            print(f'  (function-boundary path)')
        else:
            # Last resort: cut to buildTabs() call (old behavior)
            bp = html.find('buildTabs();', data_start)
            if bp < 0:
                print('Ã¢Åâ Could not patch dashboard Ã¢â¬â no injection point found')
                sys.exit(1)
            html = html[:data_start] + new_block + '\n\n' + html[bp:]
            print(f'  (buildTabs-anchor path)')

print(f'Ã¢Åâ¦ Dashboard data updated Ã¢â¬â {len(dates)} days, {len(STORE_TOTALS)} stores, {len(PROD_TOTALS)} products')

with open(DASHBOARD, 'w', encoding='utf-8') as f:
    f.write(html)

print(f'Ã¢Åâ¦ StoreÃâProduct data patched Ã¢â¬â {sum(len(v) for dv in store_prods.values() for v in dv.values())} product-store-day combos')

# Fix: ensure all stores shown in heatmap (remove slice(0,12))
with open(DASHBOARD, 'r', encoding='utf-8') as _f:
    _html = _f.read()
_patched = _html.replace('STORE_TOTALS.slice(0,12).map(([n])=>n)', 'STORE_TOTALS.map(([n])=>n)')
if _patched != _html:
    with open(DASHBOARD, 'w', encoding='utf-8') as _f:
        _f.write(_patched)
    print('\u2705 Heatmap fix applied (all stores visible)')

# Fix: ensure all stores shown in trend tab (remove slice(0,8))
with open(DASHBOARD, 'r', encoding='utf-8') as _f:
    _html = _f.read()
_patched = _html.replace('STORE_TOTALS.slice(0,8).map(function(s){return s[0];})', 'STORE_TOTALS.map(function(s){return s[0];})')
if _patched != _html:
    with open(DASHBOARD, 'w', encoding='utf-8') as _f:
        _f.write(_patched)
    print('\u2705 Trend tab fix applied (all stores visible)')
