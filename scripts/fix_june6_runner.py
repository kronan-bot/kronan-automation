"""
Recovery script: removes wrong June 6, 2026 data from master, re-processes with corrected file.
Usage: KRONAN_BASE=data python scripts/fix_june6_runner.py fix/kronan_june6_correct.xlsx
"""
import sys, os, datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from kronan_master import read_report, append_daily, rebuild_monthly, _try_parse_date, run
from openpyxl import load_workbook

_BASE = os.environ.get('KRONAN_BASE')
MASTER = os.path.join(_BASE if _BASE else '.', 'Krónan_Master_Skrá.xlsx')
TARGET_DATE = datetime.date(2026, 6, 6)

def remove_date_from_daily(wb, target_date):
    total_removed = 0
    for sname in ["Dagleg - Verslanir", "Dagleg - Vörur", "Dagleg - Vara×Verslun"]:
        if sname not in wb.sheetnames:
            print(f"  WARNING: sheet {sname!r} not found")
            continue
        ws = wb[sname]
        to_delete = []
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if row[0] is None:
                continue
            d = _try_parse_date(row[0])
            if d:
                ddate = d.date() if hasattr(d, 'date') else d
                if ddate == target_date:
                    to_delete.append(i)
        for idx in reversed(to_delete):
            ws.delete_rows(idx)
        total_removed += len(to_delete)
        print(f"  {sname}: removed {len(to_delete)} rows for {target_date}")
    return total_removed

def main():
    corrected_file = sys.argv[1] if len(sys.argv) > 1 else None
    if not corrected_file or not os.path.exists(corrected_file):
        print(f"ERROR: corrected file not found: {corrected_file!r}")
        sys.exit(1)

    print(f"\n📂 Loading master: {MASTER}")
    wb = load_workbook(MASTER)
    print(f"   Sheets: {wb.sheetnames}")

    print(f"\n🗑  Removing wrong {TARGET_DATE} data from daily sheets...")
    n = remove_date_from_daily(wb, TARGET_DATE)
    print(f"   Total removed: {n} rows")

    print(f"\n💾 Saving cleaned master...")
    wb.save(MASTER)
    print("   Saved.")

    print(f"\n📥 Processing corrected file: {corrected_file}")
    run(corrected_file)

    print(f"\n✅ Fix complete.")

if __name__ == "__main__":
    main()
